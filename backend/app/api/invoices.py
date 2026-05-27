import asyncio
import csv
import io
import re
import uuid
from pathlib import Path
from typing import Any

from fastapi import APIRouter, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app import config
from app import db
from app.services.analyzer import CATEGORIES, analyze_invoice_text
from app.services.bielik import analyze_with_bielik, merge_invoice_analysis
from app.services.ocr import extract_text, tesseract_available

router = APIRouter(prefix="/invoices", tags=["invoices"])

_ALLOWED_MIME = frozenset(
    {
        "application/pdf",
        "application/octet-stream",
        "image/jpeg",
        "image/png",
        "image/jpg",
    }
)

# Ordered columns for CSV / XLSX export
_EXPORT_FIELDS: list[tuple[str, str]] = [
    ("id", "ID dokumentu"),
    ("original_filename", "Nazwa pliku"),
    ("status", "Status"),
    ("created_at", "Dodano"),
    ("invoice_number", "Numer faktury"),
    ("issue_date", "Data wystawienia"),
    ("sale_date", "Data sprzedaży"),
    ("payment_due_date", "Termin zapłaty"),
    ("issue_place", "Miejsce wystawienia"),
    ("seller_name", "Nazwa sprzedawcy"),
    ("seller_nip", "NIP sprzedawcy"),
    ("buyer_name", "Nazwa nabywcy"),
    ("buyer_nip", "NIP nabywcy"),
    ("net_amount", "Suma netto"),
    ("vat_amount", "Suma VAT"),
    ("gross_amount", "Suma brutto"),
    ("currency", "Waluta"),
    ("category", "Kategoria"),
    ("confirmed", "Zatwierdzone"),
    ("extraction_method", "Metoda ekstrakcji"),
    ("bielik_status", "Status Bielik"),
]


# ---------------------------------------------------------------------------
# Pydantic validation model for PATCH /analysis
# ---------------------------------------------------------------------------

class AnalysisPatch(BaseModel):
    """Typed, validated merge-patch for invoice analysis fields."""

    model_config = {"extra": "ignore"}

    invoice_number: str | None = None
    issue_date: str | None = None
    sale_date: str | None = None
    payment_due_date: str | None = None
    issue_place: str | None = None
    seller_nip: str | None = None
    buyer_nip: str | None = None
    seller_name: str | None = None
    buyer_name: str | None = None
    net_amount: str | None = None
    vat_amount: str | None = None
    gross_amount: str | None = None
    currency: str | None = None
    category: str | None = None
    confirmed: bool | None = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_stem(name: str, max_len: int = 120) -> str:
    base = Path(name).name
    base = re.sub(r"[^\w.\-]", "_", base, flags=re.UNICODE)
    if not base or base.startswith("."):
        base = "document"
    return base[:max_len]


def _extension(filename: str | None) -> str:
    if not filename:
        return ""
    return Path(filename).suffix.lower()


def _flatten_invoice(item: dict) -> dict[str, Any]:
    """Merge top-level invoice fields with analysis fields for export."""
    analysis = item.get("analysis") or {}
    top_keys = {"id", "original_filename", "status", "created_at"}
    return {
        "id": item.get("id", ""),
        "original_filename": item.get("original_filename", ""),
        "status": item.get("status", ""),
        "created_at": item.get("created_at", ""),
        **{k: analysis.get(k, "") for k, _ in _EXPORT_FIELDS if k not in top_keys},
    }


# ---------------------------------------------------------------------------
# Upload / process endpoints
# ---------------------------------------------------------------------------

@router.post(
    "/upload",
    status_code=status.HTTP_201_CREATED,
    summary="Prześlij fakturę (JPG, PNG, PDF)",
)
async def upload_invoice(file: UploadFile = File(...)) -> dict:
    ext = _extension(file.filename)
    if ext not in config.ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Dozwolone rozszerzenia: {', '.join(sorted(config.ALLOWED_EXTENSIONS))}. "
                f"Otrzymano: {ext or '(brak)'}"
            ),
        )

    content_type = (file.content_type or "").split(";")[0].strip().lower()
    if content_type and content_type not in _ALLOWED_MIME:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Nieobsługiwany typ MIME: {content_type}",
        )

    doc_id = str(uuid.uuid4())
    stem = _safe_stem(file.filename or "document")
    stored_name = f"{doc_id}_{stem}"
    dest = config.UPLOAD_DIR / stored_name

    config.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    size = 0
    try:
        with dest.open("wb") as out:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                if size > config.MAX_UPLOAD_BYTES:
                    raise HTTPException(
                        status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                        detail=f"Plik przekracza limit {config.MAX_UPLOAD_BYTES} bajtów.",
                    )
                out.write(chunk)
    except HTTPException:
        if dest.exists():
            dest.unlink(missing_ok=True)
        raise
    except OSError as e:
        if dest.exists():
            dest.unlink(missing_ok=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Nie udało się zapisać pliku: {e}",
        ) from e

    # SHA-256 deduplication check
    sha256 = await asyncio.to_thread(db.compute_sha256, dest)
    existing_id = db.find_duplicate_sha256(sha256)
    if existing_id:
        dest.unlink(missing_ok=True)
        existing = db.get_invoice(existing_id) or {}
        return {
            "id": existing_id,
            "duplicate": True,
            "status": existing.get("status"),
            "original_filename": existing.get("original_filename"),
            "stored_path": existing.get("stored_path", ""),
            "stored_filename": Path(existing.get("stored_path", "")).name,
            "content_type": existing.get("content_type"),
            "size_bytes": existing.get("size_bytes", size),
        }

    db.create_invoice(
        invoice_id=doc_id,
        original_filename=file.filename,
        stored_path=dest,
        content_type=file.content_type,
        size_bytes=size,
        file_sha256=sha256,
    )

    return {
        "id": doc_id,
        "duplicate": False,
        "status": "uploaded",
        "original_filename": file.filename,
        "stored_path": str(dest),
        "stored_filename": stored_name,
        "content_type": file.content_type,
        "size_bytes": size,
    }


@router.post(
    "/upload-and-process",
    status_code=status.HTTP_201_CREATED,
    summary="Prześlij i od razu przetwórz dokument",
)
async def upload_and_process_invoice(file: UploadFile = File(...)) -> dict:
    uploaded = await upload_invoice(file)
    # If a duplicate was found return the already-processed invoice directly
    if uploaded.get("duplicate"):
        existing = db.get_invoice(uploaded["id"])
        if existing:
            existing["duplicate"] = True
            return existing
    processed = await process_invoice(uploaded["id"])
    processed["upload_status_code"] = 201
    return processed


# ---------------------------------------------------------------------------
# Static / export / utility routes — MUST appear before /{invoice_id}
# ---------------------------------------------------------------------------

@router.get("", summary="Lista dokumentów z opcjonalnym filtrowaniem")
async def list_invoices(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0, description="Numer pierwszego rekordu do pobrania (stronicowanie)"),
    category: str | None = Query(None, description="Filtruj po kategorii"),
    status: str | None = Query(None, description="Filtruj po statusie: uploaded|processing|processed|failed"),
    date_from: str | None = Query(None, description="Data wystawienia od (YYYY-MM-DD)"),
    date_to: str | None = Query(None, description="Data wystawienia do (YYYY-MM-DD)"),
    search: str | None = Query(None, description="Szukaj w nazwie pliku i numerze faktury"),
) -> list[dict]:
    return db.list_invoices(
        limit=limit,
        offset=offset,
        category=category,
        status=status,
        date_from=date_from,
        date_to=date_to,
        search=search,
    )


@router.get("/categories", summary="Lista kategorii wydatków")
async def list_categories() -> list[str]:
    """Return the ordered list of expense categories defined in analyzer.py."""
    return CATEGORIES


@router.get("/stats", summary="Statystyki wydatków (suma brutto per kategoria i miesiąc)")
async def get_stats() -> dict:
    return db.get_stats()


@router.get("/status/ocr", summary="Status OCR")
async def ocr_status() -> dict:
    return {
        "tesseract_available": tesseract_available(),
        "supported_file_types": sorted(config.ALLOWED_EXTENSIONS),
        "ocr_lang": config.OCR_LANG,
    }


@router.get("/export/csv", summary="Eksportuj faktury do CSV")
async def export_csv(
    category: str | None = Query(None),
    status: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    search: str | None = Query(None),
    limit: int = Query(200, ge=1, le=1000),
) -> StreamingResponse:
    invoices = db.list_invoices(
        limit=limit,
        category=category,
        status=status,
        date_from=date_from,
        date_to=date_to,
        search=search,
    )
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([label for _, label in _EXPORT_FIELDS])
    for inv in invoices:
        flat = _flatten_invoice(inv)
        writer.writerow([flat.get(field, "") for field, _ in _EXPORT_FIELDS])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=faktury.csv"},
    )


@router.get("/export/xlsx", summary="Eksportuj faktury do XLSX")
async def export_xlsx(
    category: str | None = Query(None),
    status: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    search: str | None = Query(None),
    limit: int = Query(200, ge=1, le=1000),
) -> StreamingResponse:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Biblioteka openpyxl nie jest zainstalowana.",
        ) from exc

    invoices = db.list_invoices(
        limit=limit,
        category=category,
        status=status,
        date_from=date_from,
        date_to=date_to,
        search=search,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faktury"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    headers = [label for _, label in _EXPORT_FIELDS]
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(16, len(headers[col_idx - 1]) + 4)

    for inv in invoices:
        flat = _flatten_invoice(inv)
        ws.append([flat.get(field, "") for field, _ in _EXPORT_FIELDS])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=faktury.xlsx"},
    )


# ---------------------------------------------------------------------------
# Per-invoice operations — parametric routes last to avoid shadowing
# ---------------------------------------------------------------------------

@router.post(
    "/{invoice_id}/process",
    summary="Przetwórz dokument (OCR + analiza pól)",
)
async def process_invoice(invoice_id: str) -> dict:
    invoice = db.get_invoice(invoice_id)
    if invoice is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Nie znaleziono dokumentu.")

    file_path = Path(invoice["stored_path"])
    if not file_path.exists():
        db.mark_failed(invoice_id, "Brak pliku źródłowego na dysku.")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Plik źródłowy nie istnieje.",
        )

    db.mark_processing(invoice_id)

    def _run_pipeline() -> tuple[str, dict]:
        """Synchronous OCR + analysis pipeline — runs in a thread pool."""
        ocr_result = extract_text(file_path)
        text = ocr_result["text"]
        heuristic = analyze_invoice_text(text)
        bielik = analyze_with_bielik(text, heuristic)
        analysis = merge_invoice_analysis(heuristic, bielik)
        analysis["ocr_engine"] = ocr_result["engine"]
        analysis["ocr_warning"] = ocr_result["warning"]
        analysis["ocr_text_length"] = len(text)
        db.mark_processed(invoice_id, ocr_text=text, analysis=analysis)
        return text, analysis

    try:
        text, analysis = await asyncio.to_thread(_run_pipeline)
    except Exception as exc:  # noqa: BLE001
        db.mark_failed(invoice_id, str(exc))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Błąd przetwarzania dokumentu: {exc}",
        ) from exc

    response = db.get_invoice(invoice_id) or {"id": invoice_id, "status": "processed"}
    preview = (text[:220] + "...") if len(text) > 220 else text
    response["processing_summary"] = {
        "engine": analysis.get("ocr_engine"),
        "text_length": len(text),
        "preview": preview,
        "warning": analysis.get("ocr_warning"),
    }
    return response


@router.get("/{invoice_id}/events", summary="Historia operacji dla dokumentu")
async def get_invoice_events(invoice_id: str) -> list[dict]:
    if db.get_invoice(invoice_id) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Nie znaleziono dokumentu.")
    return db.get_invoice_events(invoice_id)


@router.get("/{invoice_id}", summary="Pobierz szczegóły dokumentu")
async def get_invoice(invoice_id: str) -> dict:
    invoice = db.get_invoice(invoice_id)
    if invoice is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Nie znaleziono dokumentu.")
    return invoice


@router.patch(
    "/{invoice_id}/analysis",
    summary="Zaktualizuj (nadpisz) wybrane pola analizy faktury",
)
async def update_invoice_analysis(invoice_id: str, patch: AnalysisPatch) -> dict:
    patch_dict = patch.model_dump(exclude_unset=True)
    if not patch_dict:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Brak pól do aktualizacji.",
        )
    found = db.update_invoice_analysis(invoice_id, patch_dict)
    if not found:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Nie znaleziono dokumentu.")
    updated = db.get_invoice(invoice_id)
    return updated or {"id": invoice_id}
